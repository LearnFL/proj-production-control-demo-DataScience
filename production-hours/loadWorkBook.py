from openpyxl import load_workbook
from openpyxl.styles import Font
import math 

class WorkBook:
    def _load(self, path):
        self.wb = load_workbook(path)
        self.ws1 = self.wb['Schedule']

        try:
            self.ws2 = self.wb['Pipeline Hours'] 
        except:
           self.ws2 = self.wb.create_sheet("Pipeline Hours")
        
        try:
            self.ws3 = self.wb["Total Per Job Type"] 
        except:
           self.ws3 = self.wb.create_sheet("Total Per Job Type")
    
    def prepWorkBook(self, cellRange=['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1'], 
                                machinesList=['TOTAL', 'MACHINE_1', 'MACHINE_2', 'MACHINE_3', 'MACHINE_4', 'MACHINE_5', 'MACHINE_6']):
        self._load(self.path)

        for cell, mach in zip(cellRange, machinesList):
            self.ws2[cell] = mach
            self.ws2[cell].font = Font(bold=True)

    def _machineHoursBreakdown(self, total_hours):
        self.ws2['A2'] = total_hours
        self.ws2['B2'] = self.ml_Hours
        self.ws2['C2'] = self.ts4_Hours
        self.ws2['D2'] = self.sw_Hours
        self.ws2['E2'] = self.sl_Hours
        self.ws2['F2'] = self.lap_Hours
        self.ws2['G2'] = self.rnd_Hours

        self.ws2['A3'] = 'WEEKS'
        self.ws2['A3'].font = Font(bold=True)
        self.ws2['B3'] = math.ceil(self.ml_Hours / 80)
        self.ws2['C3'] = math.ceil(self.ts_Hours / 240)
        self.ws2['D3'] = math.ceil(self.sw_Hours / 160)
        self.ws2['E3'] = math.ceil(self.sl_Hours / 240)
        self.ws2['F3'] = math.ceil(self.lap_Hours / 80)
        self.ws2['G3'] = math.ceil(self.rnd_Hours / 80)
        
        self._machineTimePerMonth()

        self.wb.save(self.path) 

    def _machineTimePerMonth(self):
        colN = 1
        rowN = 5
        monthList = ['','January','February','March','April','May','June','July','August','September','October','November','December']
        for month, machine in self.log.items():
            self.ws2.cell(row=rowN, column=colN).value = monthList[math.trunc(month)]
            self.ws2.cell(row=rowN, column=colN).font = Font(bold=True)
           
            for mach, hours in machine.items():
                colN += 1
                self.ws2.cell(row=rowN, column=colN).value = mach
                self.ws2.cell(row=rowN, column=colN).font = Font(color="FF000080")
                rowN += 1
                self.ws2.cell(row=rowN, column=colN).value = hours
                rowN -= 1 
                
            rowN += 2
            colN = 1
    
       def _totalPerJobType(self, data):
            for i, r in data.iteritems():
                self.ws3.append([i, r])
            self.wb.save(self.path) 
